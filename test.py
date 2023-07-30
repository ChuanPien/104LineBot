# import jieba
# jieba.set_dictionary('dict.txt.big')
# jieba.initialize()  #強制提早加載

# import paddle
# paddle.enable_static()
# jieba.enable_paddle()# 启动paddle模式。 0.40版之后开始支持，早期版本不支持
# strs=["我来到北京清华大学","乒乓球拍卖完了","中国科学技术大学"]
# for str in strs:
#     seg_list = jieba.cut(str,use_paddle=True) # 使用paddle模式
#     print("Paddle Mode: " + '/'.join(list(seg_list)))

# 全模式
# seg_list = jieba.cut("我来到北京清华大学", cut_all=True)
# print("\nFull Mode: " + "/ ".join(seg_list))

# fix = ['軟體工程師', '沒有工作經驗']
# for i in fix:
#     jieba.suggest_freq((i), True)  #可調節單個詞語的詞頻，使其能（或不能）被分出來

# seg_list = jieba.cut("我想找軟體工程師無學歷沒有工作經驗")  # 默认是精确模式
# print(", ".join(seg_list))

# seg_list = jieba.cut_for_search("小明硕士毕业于中国科学院计算所，后在日本京都大学深造")  # 搜索引擎模式
# print("\n, ".join(seg_list))

# ------------------------------------------------------

# https://www.tpisoftware.com/tpu/articleDetails/2013
# from sklearn.feature_extraction.text import CountVectorizer
# from sklearn.linear_model import LogisticRegression
# import pandas as pd

# read = pd.read_excel("Crawler/data.xlsx").values.tolist()
# corpus = [row[0] for row in read]
# intents = [row[1] for row in read]

# feature_extractor = CountVectorizer(
#             analyzer="word", ngram_range=(1, 2), binary=True,
#             token_pattern=r'([a-zA-Z]+|\w)')
# X = feature_extractor.fit_transform(corpus)

# INTENT_CLASSIFY_REGULARIZATION = "l2"

# lr = LogisticRegression(penalty=INTENT_CLASSIFY_REGULARIZATION,class_weight='balanced')
# lr.fit(X, intents)

# user_input = ['松山']
# X2 = feature_extractor.transform(user_input)
# print(lr.predict(X2))

# probs = lr.predict_proba(X2)[0]
# for predict_intent, prob in sorted(zip(lr.classes_, probs), key = lambda x: x[1],reverse = True):
#     print(predict_intent, prob)

# ------------------------------------------------------
from openpyxl import Workbook, load_workbook

wb = load_workbook("Crawler/data.xlsx")
ws = wb["city"]

A = "東區"

for row in range(1, ws.max_row):
    produceName = ws.cell(row, 1).value
    if A in produceName:
        print(ws.cell(row, 2).value + "|" + A)
