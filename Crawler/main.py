#import selenium 相關函式庫
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook, load_workbook
import time, json, os

#讀取設定檔
with open('UserConfig/config.json', 'r', encoding='utf8')as J:
    j = json.load(J)

#寫入JSON
def write_json(new_data):                 
    file = 'Crawler/city1.json'                                      #city.json地區資料/job.json工作資料
    if not os.path.isfile(file):
        with open(file, "w") as f:
            json.dump([], f)
            f.close()

    with open(file,'r+', encoding='utf8') as f:
        file_data = json.load(f)                                                #讀取JSON現有資料
        file_data.append(new_data)                                              #將新資料與舊資料結合
        f.seek(0)                                                               #設定輸入位置
        f.write(json.dumps(file_data, ensure_ascii = False, indent = 2))        #寫入                                                   
        f.close()                                                               #關閉

'''存入excel才需啟用(9、80、89行也是)
# wb = Workbook()                           #建立空白excel，最後存檔時才命名
wb = load_workbook('data.xlsx')             #開啟excel
wb.create_sheet("job")                      #建立工作表
ws = wb['job']                              #選擇工作表
'''

#設定Chrome參數
def setup():
    s = Service(j['driver'])
    options = Options()
    options.add_experimental_option('useAutomationExtension', False)
    options.add_experimental_option("excludeSwitches", ['enable-automation'])
    global driver
    driver = webdriver.Chrome(service=s, options=options)
    # driver.maximize_window()                                          #全螢幕
    driver.get("https://www.104.com.tw/jobs/main/")                     #連到104網頁

setup()

#等待網頁元素出現
WebDriverWait(driver, 100.0).until(
    EC.presence_of_element_located((By.ID, "ijob"))
    )

#模擬點擊[職務類別選單]
ijob = driver.find_element(By.ID, "ijob")                          #icity資料選單/ijob資料選單
action = ActionChains(driver).click(ijob).perform()

#等待網頁元素出現
WebDriverWait(driver, 100.0).until(
    EC.presence_of_element_located((By.CLASS_NAME, "arrow--right"))
    )

#抓取所有職務類別並存入JSON檔
arrows = driver.find_elements(By.CLASS_NAME, "arrow--right")                        #抓取大類別
for arrow in arrows:
    i = 0
    data = []
    action = ActionChains(driver).click(arrow).perform()                            #依序打開
    ardowns = driver.find_elements(By.CLASS_NAME, "arrow--down")                    #抓取中類別
    for ardown in ardowns[1:]:
        action = ActionChains(driver).click(ardown).perform()                       #依序打開
        l2s = driver.find_elements(By.CLASS_NAME, "category-item--level-two")       #抓取中類別
        l3s = driver.find_elements(By.CLASS_NAME, "category-item--level-three")     #抓取小類別
        time.sleep(0.2)                                                             #等待0.2秒
        #依序將l3s中的文字提出並放入data陣列中
        for l3 in l3s:
            if l3.text != '':
                data.append(l3.text)
                # ws.append([l3.text,l2s[i].text])                                     #寫入excel
        write_json({l2s[i].text: data,})                                            #呼叫寫入JSON方法
        l3s.clear()                                                                  #清空l3s資料
        data.clear()                                                                 #清空data資料
        i += 1
        arrowup = driver.find_element(By.CLASS_NAME, "arrow--up")                   #抓取打開的小類別
        actions = ActionChains(driver).click(arrowup).perform()                     #關閉小類別
        time.sleep(0.3)                                                             #等待0.3秒

# wb.save('data.xlsx')                                                                #excel存檔
driver.quit()                                                                       #關閉網頁